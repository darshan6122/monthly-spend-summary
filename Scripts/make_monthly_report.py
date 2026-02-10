#!/usr/bin/env python3
"""
Build a clear, easy-to-scan monthly report from your CIBC CSV exports.
Transactions sheet is the source of truth: change a Category there and
Summary + By Category update automatically (Excel formulas).

Sheets:
  1. Transactions – All transactions; edit Category here to recategorize
  2. Summary       – Formulas: Income, Expenses (excl. transfers), Net, spending by category
  3. By Category   – Formulas: same category breakdown (updates when you change Transactions)

Totals exclude "Transfers & Payments" (e.g. credit card payments, internal transfers)
so Income and Expenses show believable numbers.

Usage:
  .venv/bin/python make_monthly_report.py              # asks which folder
  .venv/bin/python make_monthly_report.py "DECEMBER 2025"
  (App sets EXPENSE_REPORTS_ACCOUNTS_DIR and passes folder name.)
"""

import csv
import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

# App can set EXPENSE_REPORTS_ACCOUNTS_DIR so the script uses that folder explicitly.
_raw = os.environ.get("EXPENSE_REPORTS_ACCOUNTS_DIR")
ACCOUNTS_DIR = Path(_raw).resolve() if _raw else Path(__file__).resolve().parent

# Built-in rules (used if category_rules.json is not present in Accounts folder).
_BUILTIN_CATEGORY_RULES = [
    # Work income (only this counts as "work" — shown separately in Summary)
    (r"(?i)electronic funds transfer pay windreg|pay windreg", "Work Income"),
    # Other money in = from savings accounts or from friends (excluded from main Income total)
    (r"(?i)payment thank you|paiemen t merci|internet transfer 0{6,}|interac transfer", "Transfers & Payments"),
    (r"(?i)e-transfer.*darshan bodara|e-transfer.*dev shah|e-transfer.*sumitkumar", "Transfers & Payments"),
    (r"(?i)internet banking (e-transfer|internet transfer)", "Transfers & Payments"),
    # Bills & utilities
    (r"(?i)rogers \*|rogers\*\*\*\*\*\*", "Subscriptions & Bills"),
    (r"(?i)enwin|university of windsor|bill pay", "Utilities & Bills"),
    (r"(?i)apple\.com|cursor.*cursor\.com|paypal", "Subscriptions & Bills"),
    # Food & drink
    (r"(?i)tim hortons|starbucks|mcdonald|presotea|taco bell|subway|pizza pizza|chipotle|burger king|new york fries|dollarama|miniso", "Food & Drink"),
    (r"(?i)athidhi|janpath|spago|chilly bliss|paan banaras|restaurant", "Restaurants"),
    # Shopping & groceries
    (r"(?i)instacart|costco|wal-mart|amzn |amazon\.ca|temu", "Shopping & Groceries"),
    # Transport
    (r"(?i)uber|lyft|vets cab|presto fare|pearson parking|michigan flyer|spirit air|air can", "Transport & Travel"),
    # Other
    (r"(?i)sport chek|cinplex|vue", "Entertainment"),
    (r"(?i)shell |gas |petrol", "Gas & Auto"),
    (r"(?i)interest|service charge|fee", "Fees & Interest"),
    (r"(?i)chiropractic", "Health"),
    (r"(?i)shoppers drug|pharmacy", "Pharmacy"),
    (r"(?i)sephora", "Personal Care"),
    (r"(?i)branch transaction|automated banking machine", "Fees & Interest"),
]
# Fallback: catch remaining transfer-like (don't match "internet banking bill pay")
TRANSFER_FALLBACK = (r"(?i)e-transfer|internet transfer\s|interac\s+transfer", "Transfers & Payments")

_BUILTIN_ALL_CATEGORIES = [
    "Work Income", "Transfers & Payments", "Shopping & Groceries", "Food & Drink",
    "Restaurants", "Transport & Travel", "Subscriptions & Bills", "Utilities & Bills",
    "Entertainment", "Fees & Interest", "Health", "Pharmacy", "Personal Care", "Gas & Auto",
    "Uncategorized",
]


def _load_category_config() -> tuple[list, list]:
    """Load (rules, all_categories) from ACCOUNTS_DIR/category_rules.json if present; else built-in."""
    path = ACCOUNTS_DIR / "category_rules.json"
    if not path.exists():
        return (_BUILTIN_CATEGORY_RULES, _BUILTIN_ALL_CATEGORIES)
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        rules = []
        for r in data.get("rules", []):
            if isinstance(r, dict) and r.get("pattern") and r.get("category"):
                rules.append((str(r["pattern"]), str(r["category"])))
        categories = list(data.get("categories", []))
        return (rules if rules else _BUILTIN_CATEGORY_RULES, categories if categories else _BUILTIN_ALL_CATEGORIES)
    except (json.JSONDecodeError, OSError):
        return (_BUILTIN_CATEGORY_RULES, _BUILTIN_ALL_CATEGORIES)


# Defaults; run() overwrites from category_rules.json when present.
CATEGORY_RULES = _BUILTIN_CATEGORY_RULES
ALL_CATEGORIES = list(_BUILTIN_ALL_CATEGORIES)


def suggest_category(description: str) -> str:
    if not (description and description.strip()):
        return "Uncategorized"
    for pattern, category in CATEGORY_RULES:
        if re.search(pattern, description):
            return category
    # Catch generic transfers so they don't inflate Income/Expenses
    if re.search(TRANSFER_FALLBACK[0], description):
        return TRANSFER_FALLBACK[1]
    return "Uncategorized"


def parse_amount(s: str) -> float:
    if not s or not s.strip():
        return 0.0
    try:
        return float(s.strip().replace(",", ""))
    except ValueError:
        return 0.0


def friendly_account(name: str) -> str:
    """Turn 'cibc 2682' into something readable."""
    if "chq" in name.lower():
        return "Chequing"
    if "sav" in name.lower():
        return "Savings"
    if "2682" in name:
        return "Credit (2682)"
    if "0853" in name:
        return "Credit (0853)"
    return name


def read_cibc_csv(filepath: Path) -> list[dict]:
    rows = []
    source_name = filepath.stem
    with open(filepath, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 2:
                continue
            date = (row[0] or "").strip()
            desc = (row[1] or "").strip()
            debit = parse_amount(row[2] if len(row) > 2 else "")
            credit = parse_amount(row[3] if len(row) > 3 else "")
            # One amount: positive = money in, negative = money out
            amount = credit - debit if (credit or debit) else 0
            category = suggest_category(desc)
            rows.append({
                "Date": date,
                "Description": desc,
                "Account": friendly_account(source_name),
                "Category": category,
                "Amount": amount,
                "Debit": debit,
                "Credit": credit,
            })
    return rows


def get_month_folders() -> list[Path]:
    month_folders = []
    for item in ACCOUNTS_DIR.iterdir():
        if item.is_dir() and not item.name.startswith("."):
            if any(item.glob("cibc*.csv")):
                month_folders.append(item)
    return sorted(month_folders, key=lambda p: (p.name.split()[-1], p.name))


# Dark-mode friendly: light gray fill + dark text so Excel Dark Mode looks good
def style_header(cell):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    cell.font = Font(bold=True, color="1F2933")
    cell.fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_currency(cell):
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    cell.font = Font(color="1F2933")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_date_cell(cell):
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    cell.font = Font(color="1F2933")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_subtotal_cell(cell, bold=True):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    if bold:
        cell.font = Font(bold=True, color="1F2933")
    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.number_format = '"$"#,##0.00'


def run():
    global CATEGORY_RULES, ALL_CATEGORIES
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        print("Need openpyxl. Run:  .venv/bin/pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    CATEGORY_RULES, ALL_CATEGORIES = _load_category_config()

    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if args:
        folder_arg = args[0].strip()
        if not folder_arg:
            print("Error: Month folder name cannot be empty.", file=sys.stderr)
            sys.exit(1)
        month_path = ACCOUNTS_DIR / folder_arg
    else:
        month_folders = get_month_folders()
        if not month_folders:
            print("No month folders with cibc*.csv found.", file=sys.stderr)
            sys.exit(1)
        print("Available folders:")
        for i, p in enumerate(month_folders, 1):
            print(f"  {i}. {p.name}")
        folder_name = input("\nWhich folder? (name or number): ").strip()
        if not folder_name:
            sys.exit(0)
        if folder_name.isdigit():
            idx = int(folder_name)
            month_path = month_folders[idx - 1] if 1 <= idx <= len(month_folders) else month_folders[0]
        else:
            month_path = ACCOUNTS_DIR / folder_name

    if not month_path.is_dir():
        print(f"Folder not found: {month_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Report folder: {month_path.name}", file=sys.stderr)
    csv_files = sorted(month_path.glob("cibc*.csv"))
    use_merged = os.environ.get("USE_MERGED_CATEGORIES", "").strip().lower() in ("1", "true", "yes")
    combined_path = month_path / f"{month_path.name}_combined.csv"

    # Load transaction_splits.json: expand any row with splits into multiple rows
    splits_map = {}
    splits_path = month_path / "transaction_splits.json"
    if splits_path.exists():
        try:
            with open(splits_path, encoding="utf-8") as sf:
                splits_map = json.load(sf)
        except (json.JSONDecodeError, OSError):
            pass

    def expand_splits(rows):
        out = []
        for r in rows:
            key = f"{r['Date']}|{r['Description']}|{r['Amount']}"
            key_alt = f"{r['Date']}|{r['Description']}|{int(r['Amount'])}"
            parts = splits_map.get(key) or splits_map.get(key_alt)
            if parts and isinstance(parts, list) and len(parts) > 0:
                for part in parts:
                    cat = part.get("category") or "Uncategorized"
                    amt = float(part.get("amount", 0))
                    if r["Amount"] < 0:
                        amt = -abs(amt)
                    out.append({
                        "Date": r["Date"],
                        "Description": r["Description"],
                        "Account": r["Account"],
                        "Category": cat,
                        "Amount": amt,
                    })
            else:
                out.append(r)
        return out

    if use_merged and combined_path.exists():
        all_rows = []
        with open(combined_path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                debit = parse_amount(row.get("Debit", "") or "0")
                credit = parse_amount(row.get("Credit", "") or "0")
                amount = credit - debit
                all_rows.append({
                    "Date": (row.get("Date") or "").strip(),
                    "Description": (row.get("Description") or "").strip(),
                    "Account": (row.get("Source") or row.get("Account", "")).strip() or "—",
                    "Category": (row.get("Suggested Category") or "").strip() or "Uncategorized",
                    "Amount": amount,
                })
        all_rows = expand_splits(all_rows)
        all_rows.sort(key=lambda r: (r["Date"], r["Description"]))
        print("Using categories from merge (_combined.csv).")
    else:
        if not csv_files:
            print("No cibc*.csv in that folder.", file=sys.stderr)
            sys.exit(1)
        all_rows = []
        for f in csv_files:
            all_rows.extend(read_cibc_csv(f))
        all_rows = expand_splits(all_rows)
        all_rows.sort(key=lambda r: (r["Date"], r["Description"]))

    # Fixed range for formulas (so changing Transactions updates Summary & By Category)
    MAX_ROW = 2000
    TX = "Transactions"  # sheet name for formula refs

    template_path = ACCOUNTS_DIR / "template.xlsx"
    if template_path.exists():
        try:
            wb = openpyxl.load_workbook(template_path)
            for name in ["Transactions", "Summary", "By Category"]:
                if name in wb.sheetnames:
                    del wb[name]
            ws_tx = wb.create_sheet("Transactions", 0)
        except Exception:
            wb = openpyxl.Workbook()
            ws_tx = wb.active
            ws_tx.title = "Transactions"
    else:
        wb = openpyxl.Workbook()
        ws_tx = wb.active
        ws_tx.title = "Transactions"

    # ----- Sheet 1: Transactions (source of truth; edit Category here) -----
    tx_headers = ["Date", "Description", "Account", "Category", "Amount"]
    for c, h in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=c, value=h)
        style_header(cell)
    for i, r in enumerate(all_rows, 2):
        ws_tx.cell(row=i, column=1, value=r["Date"])
        ws_tx.cell(row=i, column=2, value=(r["Description"] or "")[:80])
        ws_tx.cell(row=i, column=3, value=r["Account"])
        ws_tx.cell(row=i, column=4, value=r["Category"])
        ws_tx.cell(row=i, column=5, value=r["Amount"])
        style_date_cell(ws_tx.cell(row=i, column=1))
        style_date_cell(ws_tx.cell(row=i, column=2))
        style_date_cell(ws_tx.cell(row=i, column=3))
        style_date_cell(ws_tx.cell(row=i, column=4))
        style_currency(ws_tx.cell(row=i, column=5))
    ws_tx.column_dimensions["A"].width = 12
    ws_tx.column_dimensions["B"].width = 48
    ws_tx.column_dimensions["C"].width = 14
    ws_tx.column_dimensions["D"].width = 22
    ws_tx.column_dimensions["E"].width = 12
    ws_tx.freeze_panes = "A2"
    # Category dropdown so you pick from list (fewer typos, faster)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(ALL_CATEGORIES)}"',
        allow_blank=True,
        showDropDown=True,
    )
    dv.error = "Pick a category from the list"
    dv.errorTitle = "Category"
    ws_tx.add_data_validation(dv)
    dv.add(f"D2:D{min(MAX_ROW, len(all_rows) + 100)}")
    # AutoFilter so you can filter by Category, Account, etc.
    ws_tx.auto_filter.ref = ws_tx.dimensions
    # Red for spending (negative amount), green for money in (positive)
    last_tx_row = len(all_rows) + 1
    amt_range = f"E2:E{last_tx_row}"
    ws_tx.conditional_formatting.add(
        amt_range,
        FormulaRule(formula=["E2<0"], font=Font(color="9C0006")),
    )
    ws_tx.conditional_formatting.add(
        amt_range,
        FormulaRule(formula=["E2>0"], font=Font(color="006100")),
    )
    # Hint for user
    ws_tx.cell(row=1, column=6, value="← Pick Category from dropdown; Summary & By Category update automatically")
    ws_tx.cell(row=1, column=6).font = Font(italic=True, color="666666")

    # ----- Sheet 2: Summary (all formulas; excludes Transfers & Payments from main totals) -----
    ws_sum = wb.create_sheet("Summary", 1)
    month_title = month_path.name
    ws_sum["A1"] = month_title
    ws_sum["A1"].font = Font(size=16, bold=True)
    ws_sum["A2"] = "Edit Category in Transactions sheet — these numbers update automatically."
    ws_sum["A2"].font = Font(italic=True, color="444444")
    ws_sum["A3"] = "Work Income = PAY WINDREG only. Other money in = savings / from friends (excl. from Income total)."
    ws_sum["A3"].font = Font(italic=True, color="444444")
    ws_sum["A4"] = "Tip: Open in Excel to see calculated numbers. Change Category in Transactions (use dropdown) → Summary updates."
    ws_sum["A4"].font = Font(italic=True, color="666666")
    row = 5
    ws_sum.cell(row=row, column=1, value="Work Income (PAY WINDREG)")
    ws_sum.cell(row=row, column=2, value=None)
    ws_sum.cell(row=row, column=2).value = f'=SUMIFS({TX}!$E$2:$E${MAX_ROW},{TX}!$E$2:$E${MAX_ROW},">0",{TX}!$D$2:$D${MAX_ROW},"Work Income")'
    ws_sum.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 1
    ws_sum.cell(row=row, column=1, value="Income (excl. transfers)")
    ws_sum.cell(row=row, column=2, value=None)
    ws_sum.cell(row=row, column=2).value = f'=SUMIFS({TX}!$E$2:$E${MAX_ROW},{TX}!$E$2:$E${MAX_ROW},">0",{TX}!$D$2:$D${MAX_ROW},"<>Transfers & Payments")'
    ws_sum.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 1
    ws_sum.cell(row=row, column=1, value="Expenses (excl. transfers)")
    ws_sum.cell(row=row, column=2, value=None)
    ws_sum.cell(row=row, column=2).value = f'=ABS(SUMIFS({TX}!$E$2:$E${MAX_ROW},{TX}!$E$2:$E${MAX_ROW},"<0",{TX}!$D$2:$D${MAX_ROW},"<>Transfers & Payments"))'
    ws_sum.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 1
    ws_sum.cell(row=row, column=1, value="Net (Income − Expenses)")
    ws_sum.cell(row=row, column=2, value=None)
    ws_sum.cell(row=row, column=2).value = f'=B6-B7'
    ws_sum.cell(row=row, column=2).number_format = '"$"#,##0.00'
    ws_sum.cell(row=row, column=1).font = Font(bold=True)
    ws_sum.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    ws_sum.cell(row=row, column=1, value="Transfers (net, excluded from above)")
    ws_sum.cell(row=row, column=2, value=None)
    ws_sum.cell(row=row, column=2).value = f'=SUMIF({TX}!$D$2:$D${MAX_ROW},"Transfers & Payments",{TX}!$E$2:$E${MAX_ROW})'
    ws_sum.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 2
    ws_sum.cell(row=row, column=1, value=None)
    ws_sum.cell(row=row, column=1).value = f'=COUNTA({TX}!$A$2:$A${MAX_ROW})'
    ws_sum.cell(row=row, column=2, value="transaction(s) this month")
    ws_sum.cell(row=row, column=1).font = Font(italic=True)
    row += 2
    ws_sum.cell(row=row, column=1, value="Spending by category (formulas)")
    ws_sum.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    headers = ["Category", "Spent", "% of spending"]
    for c, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=row, column=c, value=h)
        style_header(cell)
    row += 1
    # Total expenses (excl. transfers) for % column — in B7; reference B7 in formulas
    spending_cats = sorted(set(r["Category"] for r in all_rows if r["Amount"] < 0 and r["Category"] != "Transfers & Payments"), key=lambda k: -sum(-r["Amount"] for r in all_rows if r["Category"] == k and r["Amount"] < 0))
    if not spending_cats:
        spending_cats = ["Uncategorized"]
    start_cat_row = row
    for i, cat in enumerate(spending_cats):
        rn = start_cat_row + i
        ws_sum.cell(row=rn, column=1, value=cat)
        ws_sum.cell(row=rn, column=2, value=None)
        ws_sum.cell(row=rn, column=2).value = f'=IF(A{rn}="","",ABS(SUMIFS({TX}!$E$2:$E${MAX_ROW},{TX}!$D$2:$D${MAX_ROW},A{rn},{TX}!$E$2:$E${MAX_ROW},"<0")))'
        ws_sum.cell(row=rn, column=2).number_format = '"$"#,##0.00'
        ws_sum.cell(row=rn, column=3, value=None)
        ws_sum.cell(row=rn, column=3).value = f'=IF(B{rn}=0,"",B{rn}/B$7)'
        ws_sum.cell(row=rn, column=3).number_format = "0.0%"
        style_date_cell(ws_sum.cell(row=rn, column=1))
        style_currency(ws_sum.cell(row=rn, column=2))
        style_date_cell(ws_sum.cell(row=rn, column=3))
    other_row = start_cat_row + len(spending_cats)
    ws_sum.cell(row=other_row, column=1, value="Other (new categories you add)")
    ws_sum.cell(row=other_row, column=2, value=None)
    ws_sum.cell(row=other_row, column=2).value = f'=MAX(0,B7-SUM(B{start_cat_row}:B{other_row-1}))'
    ws_sum.cell(row=other_row, column=2).number_format = '"$"#,##0.00'
    ws_sum.cell(row=other_row, column=3, value=None)
    ws_sum.cell(row=other_row, column=3).value = f'=IF(B{other_row}=0,"",B{other_row}/B$7)'
    ws_sum.cell(row=other_row, column=3).number_format = "0.0%"
    style_date_cell(ws_sum.cell(row=other_row, column=1))
    style_currency(ws_sum.cell(row=other_row, column=2))
    style_date_cell(ws_sum.cell(row=other_row, column=3))
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 14

    # ----- Sheet 3: By Category (formulas) -----
    ws_cat = wb.create_sheet("By Category", 2)
    ws_cat.cell(row=1, column=1, value="Spending by category (from Transactions — change category there to update)")
    ws_cat.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws_cat.cell(row=2, column=1, value="Filter the Transactions sheet by Category to see individual transactions.")
    ws_cat.cell(row=2, column=1).font = Font(italic=True, color="444444")
    row = 4
    for c, h in enumerate(["Category", "Spent", "%"], 1):
        cell = ws_cat.cell(row=row, column=c, value=h)
        style_header(cell)
    row += 1
    start_cat_row_b = row
    for i, cat in enumerate(spending_cats):
        rn = start_cat_row_b + i
        ws_cat.cell(row=rn, column=1, value=cat)
        ws_cat.cell(row=rn, column=2, value=None)
        ws_cat.cell(row=rn, column=2).value = f'=IF(A{rn}="","",ABS(SUMIFS({TX}!$E$2:$E${MAX_ROW},{TX}!$D$2:$D${MAX_ROW},A{rn},{TX}!$E$2:$E${MAX_ROW},"<0")))'
        ws_cat.cell(row=rn, column=2).number_format = '"$"#,##0.00'
        ws_cat.cell(row=rn, column=3, value=None)
        ws_cat.cell(row=rn, column=3).value = f'=IF(B{rn}=0,"",B{rn}/\'Summary\'!B$7)'
        ws_cat.cell(row=rn, column=3).number_format = "0.0%"
        style_date_cell(ws_cat.cell(row=rn, column=1))
        style_currency(ws_cat.cell(row=rn, column=2))
        style_date_cell(ws_cat.cell(row=rn, column=3))
    rn = start_cat_row_b + len(spending_cats)
    ws_cat.cell(row=rn, column=1, value="Other")
    ws_cat.cell(row=rn, column=2, value=None)
    ws_cat.cell(row=rn, column=2).value = f'=MAX(0,\'Summary\'!B7-SUM(B{start_cat_row_b}:B{rn-1}))'
    ws_cat.cell(row=rn, column=2).number_format = '"$"#,##0.00'
    ws_cat.cell(row=rn, column=3, value=None)
    ws_cat.cell(row=rn, column=3).value = f'=IF(B{rn}=0,"",B{rn}/\'Summary\'!B$7)'
    ws_cat.cell(row=rn, column=3).number_format = "0.0%"
    style_date_cell(ws_cat.cell(row=rn, column=1))
    style_currency(ws_cat.cell(row=rn, column=2))
    style_date_cell(ws_cat.cell(row=rn, column=3))
    ws_cat.column_dimensions["A"].width = 28
    ws_cat.column_dimensions["B"].width = 14
    ws_cat.column_dimensions["C"].width = 14
    ws_cat.freeze_panes = "A4"

    out_name = f"{month_path.name}_Report.xlsx"
    out_path = month_path / out_name
    wb.save(out_path)
    print(f"Done. Saved: {out_path}")
    print("  – Transactions: edit Category here; other sheets update via formulas")
    print("  – Summary: Income/Expenses exclude transfers so numbers are realistic")
    print("  – By Category: same totals from formulas (filter Transactions for details)")

    # For app Quick Stats: write month_summary.json
    total_credits = sum(r["Amount"] for r in all_rows if r["Amount"] > 0)
    total_spent = sum(-r["Amount"] for r in all_rows if r["Amount"] < 0)
    by_category = defaultdict(float)
    for r in all_rows:
        if r["Amount"] < 0:
            by_category[r["Category"]] += abs(r["Amount"])
    summary = {
        "total_spent": round(total_spent, 2),
        "total_credits": round(total_credits, 2),
        "by_category": {k: round(v, 2) for k, v in sorted(by_category.items())},
        "savings_transfer": 0.0,
        "transaction_count": len(all_rows),
    }
    with open(month_path / "month_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    # Optional: HTML dashboard with Plotly
    try:
        import plotly.graph_objects as go
        import plotly.offline
        fig = go.Figure(data=[go.Bar(x=list(by_category.keys()), y=list(by_category.values()), marker_color="rgb(59, 130, 246)")])
        fig.update_layout(title=f"{month_path.name} — Spending by category", xaxis_title="Category", yaxis_title="Amount ($)", template="plotly_white", font=dict(size=12))
        html_path = month_path / "dashboard.html"
        fig.write_html(str(html_path), config={"displayModeBar": True})
        print(f"  – Dashboard: {html_path}")
    except ImportError:
        pass

if __name__ == "__main__":
    run()
